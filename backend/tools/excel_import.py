"""
Excel批量导入功能
"""
import io
from typing import List, Dict, Tuple
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from tools.admin_service import admin_service


class ExcelImportService:
    """Excel导入服务"""
    
    # 用户导入必填字段
    USER_REQUIRED_FIELDS = ["username", "password", "phone", "name"]
    # 用户导入可选字段
    USER_OPTIONAL_FIELDS = ["home_address", "health_info"]
    # 所有支持的字段
    USER_ALL_FIELDS = USER_REQUIRED_FIELDS + USER_OPTIONAL_FIELDS
    
    # 字段映射（中文表头 -> 英文字段）
    FIELD_MAPPING = {
        "用户名": "username",
        "密码": "password",
        "手机号": "phone",
        "姓名": "name",
        "老人姓名": "name",
        "邮箱": "email",
        "电子邮件": "email",
        "家庭住址": "home_address",
        "住址": "home_address",
        "健康信息": "health_info",
        "健康状况": "health_info",
    }
    
    async def import_users_from_excel(
        self,
        file_content: bytes,
        session: AsyncSession
    ) -> Tuple[int, int, List[str]]:
        """
        从Excel导入用户
        
        Returns:
            Tuple[成功数, 失败数, 错误列表]
        """
        success_count = 0
        fail_count = 0
        errors = []
        
        try:
            # 读取Excel
            wb = load_workbook(io.BytesIO(file_content))
            ws = wb.active
            
            # 获取表头
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # 映射表头
            field_indices = {}
            for idx, header in enumerate(headers):
                if header in self.FIELD_MAPPING:
                    field_indices[self.FIELD_MAPPING[header]] = idx
                elif header in self.USER_ALL_FIELDS:
                    field_indices[header] = idx
            
            # 检查必填字段
            missing_fields = [f for f in self.USER_REQUIRED_FIELDS if f not in field_indices]
            if missing_fields:
                return 0, 0, [f"Excel缺少必填字段: {', '.join(missing_fields)}"]
            
            # 逐行处理
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 提取数据
                    data = {}
                    for field, idx in field_indices.items():
                        data[field] = row[idx] if idx < len(row) else None
                    
                    # 验证数据
                    if not data.get("username"):
                        errors.append(f"第{row_idx}行: 用户名不能为空")
                        fail_count += 1
                        continue
                    
                    if not data.get("password"):
                        errors.append(f"第{row_idx}行: 密码不能为空")
                        fail_count += 1
                        continue
                    
                    if not data.get("phone"):
                        errors.append(f"第{row_idx}行: 手机号不能为空")
                        fail_count += 1
                        continue
                    
                    # 转换数据类型
                    data["username"] = str(data["username"]).strip()
                    data["password"] = str(data["password"])
                    data["phone"] = str(data["phone"]).strip()
                    data["name"] = str(data["name"]).strip() if data.get("name") else data["username"]
                    
                    if data.get("home_address"):
                        data["home_address"] = str(data["home_address"])
                    if data.get("health_info"):
                        data["health_info"] = str(data["health_info"])
                    
                    # 创建用户
                    success, message, elderly = await admin_service.create_user(data, session)
                    
                    if success:
                        success_count += 1
                    else:
                        errors.append(f"第{row_idx}行: {message}")
                        fail_count += 1
                
                except Exception as e:
                    errors.append(f"第{row_idx}行: 处理异常 - {str(e)}")
                    fail_count += 1
            
            return success_count, fail_count, errors
            
        except Exception as e:
            return 0, 0, [f"Excel文件解析失败: {str(e)}"]
    
    def generate_user_template(self) -> bytes:
        """生成用户导入模板"""
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "用户导入模板"
        
        # 设置表头
        headers = ["用户名", "密码", "手机号", "姓名", "邮箱", "家庭住址", "健康信息"]
        ws.append(headers)
        
        # 添加示例数据
        example = ["zhangsan", "123456", "13800138000", "张三", "zhangsan@example.com", "北京市海淀区", "高血压、糖尿病"]
        ws.append(example)
        
        # 添加说明
        ws.append([])
        ws.append(["说明："])
        ws.append(["1. 用户名、密码、手机号、姓名为必填项"])
        ws.append(["2. 用户名需唯一，不能重复"])
        ws.append(["3. 手机号需为11位数字"])
        ws.append(["4. 密码至少6位"])
        ws.append(["5. 邮箱用于接收提醒通知（可选）"])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()


# 全局实例
excel_import_service = ExcelImportService()
